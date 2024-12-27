import pymysql
from openpyxl import Workbook
from openpyxl import load_workbook

# db 전체선택
def select_all():
    conn = pymysql.connect(host='localhost',
                           user='test',
                           password='1234',
                           db='shop',
                           charset='utf8')
    try:
        with conn.cursor() as curs:
            sql="select * from accident_status"
            curs.execute(sql)
            rs = curs.fetchall()
            for row in rs:
                print(row)
    finally:
        conn.close()

# db에 넣기
def insert_db():
    conn = pymysql.connect(host='localhost',
                           user='test',
                           passwd='1234',
                           db='shop',
                           charset='utf8')
    try:
        with conn.cursor() as curs:
            sql='insert into accident_status values(%s, %s, %s, %s, %s)'
            wb = load_workbook('sea.xlsx',data_only=True)
            ws = wb['Sheet']

            iter_rows = iter(ws.rows)
            next(iter_rows)
            for row in iter_rows:
                curs.execute(sql, (row[0].value, row[1].value, row[2].value, row[8].value, row[9].value))
            conn.commit()
    finally:
        conn.close()
        wb.close()

if __name__ == "__main__":
    
    select_all()
          
