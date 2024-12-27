import sqlalchemy as db
import pandas as pd
import openpyxl
import csv
import os
import requests
import sqlite3
import pymysql

# def read_selected_columns(file_name, columns):
#     # 엑셀 파일을 열기.
#     workbook = openpyxl.load_workbook(file_name)
    
#     # 첫 번째 시트를 선택.
#     sheet = workbook.active
    
#     # 데이터를 저장할 리스트생성.
#     data = []
    
#     # 시트에서 선택된 열의 데이터 추출.
#     for row in sheet.iter_rows(min_col=1, max_col=sheet.max_column, values_only=True):
#         selected_data = [row[col-1] for col in columns]
#         data.append(selected_data)

#     return data

# def save_to_csv(data, output_file):
#     # CSV 파일을 쓰기 모드로 열기.
#     with open(output_file, 'w', newline='', encoding='UTF-8-sig') as f:
#         writer = csv.writer(f)
        
#     # 데이터를 한 행씩 CSV 파일에 쓰기.
#         for row in data:
#             writer.writerow(row)

# # 엑셀 파일의 경로 지정.
# file_name = 'sea.xlsx'

# # 원하는 열을 선택.
# columns = [1, 2, 3, 9, 10]

# # 데이터 추출.
# data = read_selected_columns(file_name, columns)
# # print(data)

# # # 출력할 CSV 파일의 이름을 지정.
# # output_file = os.path.join(os.path.expanduser('~'), 'Desktop', 'accident_status.csv')

# # # 데이터를 CSV 파일로 저장.
# # save_to_csv(data, output_file)

# # print(f"CSV 파일 저장성공 : {output_file}")

# # 데이터를 판다스로 변경
# accident_status = pd.DataFrame(data)
# # print(data)

# # CSV저장
# accident_status.to_csv("accidents_status.csv", index=False, encoding="UTF-8-sig")

# # db연결
# # mysql+pymysql://세션이름:세션비번@url:포트번호/db이름
# engine= db.create_engine(f"mysql+pymysql://test:1234@127.0.0.1:3306/shop")

# # DataFrame을 db의 accident_status 테이블에 저장
# # 이미 테이블이 존재하면 덮어쓰기 함
# accident_status.to_sql("accident_status", engine, index=False, if_exists="replace")

# # sql문을 실행하여 데이터를 읽어와서 DataFrame으로 변환
# accident_status_from_db = pd.read_sql("select * from accident_status", engine)

# # 출력
# print(accident_status_from_db)

# -------------- 수정 ----------------
 
db = pymysql.connect(host='localhost',
                    user='test',
                    passwd='1234',
                    db='shop',
                    charset='utf8')

cursor = db.cursor()
 
sql1 = "ALTER TABLE accident_status CHANGE accident_num accident_num VARCHAR(100)"
sql2 = "ALTER TABLE accident_status CHANGE ac_name ac_name VARCHAR(100)"
sql3 = "ALTER TABLE accident_status CHANGE ac_type ac_type VARCHAR(100)"
sql4 = "ALTER TABLE accident_status CHANGE ac_ton ac_ton VARCHAR(100)"
sql5 = "ALTER TABLE accident_status CHANGE ac_kind ac_kind VARCHAR(100)"

cursor.execute(sql1)
cursor.execute(sql2)
cursor.execute(sql3)
cursor.execute(sql4)
cursor.execute(sql5)
db.commit()
 
cursor.close()
db.close()
